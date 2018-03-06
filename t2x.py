#coding=utf-8
from data import (RoundData,BattleData,MatchData,Path)
from openpyxl import Workbook
from openpyxl import load_workbook

def trans(MATCH,BATTLE,ROUND,ora_path,calcu_path):
	print 'start'
	ultDic1={"Genji":[7,0,0,0],"Mccree":[6.367,0,0,0],"Pharah":[3,0,0,0],"Reaper":[3,0,0,0],"Soldier: 76":[7.4,0,0,0], \
			"Sombra":[0,0,0,0],"Tracer":[2,0,0,0],"Bastion":[9.5,0,0,0],"Hanzo":[0,0,0,0],"Junkrat":[11,0,0,0], \
			"Mei":[0,0,0,0],"Torbjorn":[12,0,0,0],"Widowmaker":[0,0,0,0],"D. Va":[3,0,0,0],"Orisa":[0,0,0,0], \
			"Reinhardt":[0,0,0,0],"Roadhog":[6.5,0,0,0],"Winston":[10,0,0,0],"Zarya":[0,0,0,0],"Ana":[0,0,0,0], \
			"Lucio":[0,0,0,0],"Mercy":[0,0,0,0],"Symmetra":[0,0,0,0],"Zenyatta":[0,0,0,0],"Doomfist":[5,0,0,0],"Moira":[8,0,0,0]}
	ultDic2={"Genji":[7,0,0,0],"Mccree":[6.367,0,0,0],"Pharah":[3,0,0,0],"Reaper":[3,0,0,0],"Soldier: 76":[7.4,0,0,0], \
			"Sombra":[0,0,0,0],"Tracer":[1,0,0,0],"Bastion":[9.5,0,0,0],"Hanzo":[0,0,0,0],"Junkrat":[11,0,0,0], \
			"Mei":[0,0,0,0],"Torbjorn":[12,0,0,0],"Widowmaker":[0,0,0,0],"D. Va":[3,0,0,0],"Orisa":[0,0,0,0], \
			"Reinhardt":[0,0,0,0],"Roadhog":[6.5,0,0,0],"Winston":[10,0,0,0],"Zarya":[0,0,0,0],"Ana":[0,0,0,0], \
			"Lucio":[0,0,0,0],"Mercy":[0,0,0,0],"Symmetra":[0,0,0,0],"Zenyatta":[0,0,0,0],"Doomfist":[5,0,0,0],"Moira":[8,0,0,0]}

	actionNum={"Genji":201,"Mccree":202,"Pharah":203,"Reaper":204,"Soldier: 76":205, \
				"Sombra":206,"Tracer":207,"Bastion":208,"Hanzo":209,"Junkrat":210, \
				"Mei":211,"Torbjorn":212,"Widowmaker":213,"D.Va":214,"Orisa":215, \
				"Reinhardt":216,"Roadhog":217,"Winston":218,"Zarya":219,"Ana":220, \
				"Lucio":221,"Mercy":222,"Symmetra":223,"Zenyatta":224,"Doomfist":225,"Moira":226,\
				"Demech":306,"Eliminate":1,"Critical":4,"Ult used":6,"Ult ready":300,\
				"Resurrect":305,"Suicide":5}
	mapdic={'Dorado':100001,'Route 66':100002,'Watchpoint: Gibraltar':100003,'Junkertown':100004,\
		 'Hollywood':100021,'Numbani':100022,"King's Row":100023,'Eichenwalde':100024,'Blizzard World':100025,\
		 'Volskaya Industries':100031,'Hanamura':100032,'Temple of Anubis':100033,'Horizon Lunar Colony':100034,\
		 'Nepal':100041,'Lijiang Tower':100042,'Ilios':100043,'Oasis':100044}

	herodic={"Doomfist":2025,"Genji":2001,"Mccree":2002,"Pharah":2003,"Reaper":2004,"Soldier: 76":2005,"Sombra":2006,\
		  "Tracer":2007,"Bastion":2008,"Hanzo":2009,"Junkrat":2010, "Mei":2011,"Torbjorn":2012,"Widowmaker":2013,\
		  "D.Va":2014,"Orisa":2015,"Reinhardt":2016,"Roadhog":2017,"Winston":2018,"Zarya":2019,"Ana":2020,\
		  "Lucio":2021,"Mercy":2022,"Moira":2026,"Symmetra":2023,"Zenyatta":2024}

	Teamdic={'DAL':{},'PHI':{},'HOU':{},'BOS':{},
			'SHD':{},'FLA':{},'NYE':{},'SFS':{},
			'VAL':{},'GLA':{},'SEO':{},'LDN':{}}
	Teamdic['DAL']={'CHIPSHAJEN':10001,'HARRYHOOK':10002,'MICKIE':10003,'XQC':10004,'SEAGULL':10005,'TAIMOU':10006,
				'CUSTA':10007,'COCCO':10008,'EFFECT':10009,'AKM':10010,'RASCAL':10011}
	Teamdic['PHI']={'JOWMEISTER':20001,'BOOMBOX':20002,'CARPE':20003,'SNILLO':20004,'FRAGI':20005,'EQO':20006,
				'SHADOWBURN':20007,'NEPTUNO':20008,'DAYFLY':20009,'HOTBA':20010,'POKO':20011,'SADO':20012}
	Teamdic['HOU']={'MUMA':30001,'BANI':30002,'CLOCKWORK':30003,'MENDOKUSAII':30004,'BOINK':30005,'LINKZR':30006,
				'SPREE':30007,'RAWKUS':30008,'JAKE':30009,'COOLMATT':30010}
	Teamdic['BOS']={'DREAMKAZPE':40001,'SNOW':40002,'GAMSU':40003,'NOTE':40004,'NEKO':40005,'AVAST':40006,
				'STRIKER':40007,'KALIOS':40008,'MISTAKES':40009,'KELLEX':40010}
	Teamdic['SHD']={'DIYA':50008,'UNDEAD':50005,'FREEFEEL':50001,'ROSHAN':50007,'ALTERING':50004,'MG':50006,
				'XUSHU':50002,'FIVEKING':50003,'GEGURI':50009,'ADO':50010,'SKY':50011,'FEARLESS':50012}
	Teamdic['FLA']={'MANNETEN':60001,'LOGIX':60002,'ZEBBOSAI':60003,'BISCHU':60004,'CWOOSH':60005,'TVIQ':60006,'ZUPPEH':60007,'ZAPPIS':60008}
	Teamdic['NYE']={'SAEBYEOLBE':70001,'MEKO':70002,'PINE':70003,'JANUS':70004,'JJONAK':70005,'MANO':70006,
				'LIBERO':70007,'ARK':70008}
	Teamdic['SFS']={'SUPER':80001,'BABYBAY':80002,'SINATRAA':80003,'SLEEPY':80004,'DANTEH':80005,'DHAK':80006,
				'NOMY':80007,'IDDQD':80008,'NEVIX':80009}
	Teamdic['VAL']={'SILKTHREAD':90001,'GRIMREALITY':90002,'KARIV':90003,'FATE':90004,'VERBO':90005,'ENVY':90006,
				'SPACE':90007,'NUMLOCKED':90008,'UNKOE':90009,'AGILITIES':90010,'SOON':90011}
	Teamdic['GLA']={'ASHER':100001,'SUREFOUR':100002,'IREMIIX':100003,'BISCHU':100004,'SHAZ':100005,'HYDRATION':100006,
				'BIGGOOSE':100007,'FISSURE':100008}
	Teamdic['SEO']={'BUNNY':110001,'MIRO':110002,'XEPHER':110003,'GIDO':110004,'WEKEED':110005,'MUNCHKIN':110006,
				'ZUNBA':110007,'KUKI':110008,'TOBI':110009,'RYUJEHONG':110010,'FLETA':110011}
	Teamdic['LDN']={'GESTURE':120001,'BIRDRING':120002,'BDOSIN':120004,'NUS':120005,'HOOREG':120006,
				'FURY':120008,'HAGOPEUN':120009,'WOOHYAL':120010,'PROFIT':120011,'CLOSER':120012}

	r=3
	timeCol=1
	actionCol=2
	subjectCol=3
	heroCol=4
	objectCol=5
	criticalCol=8

	subjectList=[]
	actionList=[]
	objectList=[]
	timeList=[]

	sNum=0
	begin=BATTLE.begin
	count=BATTLE.count
	startlist=BATTLE.startlist
	endlist=BATTLE.endlist
	winlist=BATTLE.winlist

	sbegin=ROUND.begin	
	scount=ROUND.count
	sstartlist=ROUND.startlist
	sendlist=ROUND.endlist

	sheetNum=MATCH.game

	print 'Loading...'
	wb_trans=load_workbook(ora_path,read_only=False)
	sheet_trans=wb_trans.worksheets[0]
	wb_calcu=load_workbook(calcu_path,read_only=False,keep_vba=True)
	sheet_calcu=wb_calcu.worksheets[int(sheetNum)-1]
	print 'Load completed'
	rr=1
	ca=1
	cb=2
	cc=3
	i=0
	##########################

	suicide=0
	while 1:
		time=sheet_trans.cell(row=r,column=timeCol).value #时间
		action=sheet_trans.cell(row=r,column=actionCol).value #初始事件名
		subject=sheet_trans.cell(row=r,column=subjectCol).value #对象列
		hero=sheet_trans.cell(row=r,column=heroCol).value	#英雄列
		obj=sheet_trans.cell(row=r,column=objectCol).value #目标列
		critical=sheet_trans.cell(row=r,column=criticalCol).value #暴击事件

		if time ==None:
			break

		#事件
		if action=='Hero switch': #换英雄事件
			actionList.insert(sNum,actionNum[hero])  
		if action=='Eliminate':
			if subject == None:
				suicide=1
				actionList.insert(sNum,actionNum['Suicide'])  
			else:
				if ultDic1[hero][1]==0 and ultDic2[hero][1]==0:
					if critical!="Y":
						actionList.insert(sNum,actionNum['Eliminate'])
					else:
						actionList.insert(sNum,actionNum['Critical'])

				else:
					playernum=int(subject[6:len(subject)])
					t=BATTLE.t2s(time)
					if playernum <=6:
						if playernum == ultDic1[hero][3]:
							if t <= ultDic1[hero][2]:
								actionList.insert(sNum,2)
							if t > ultDic1[hero][2]:
								ultDic1[hero][3]=0
								ultDic1[hero][1]=0
								ultDic1[hero][2]=0
					if playernum >6:
						if playernum == ultDic2[hero][3]:
							if t <= ultDic2[hero][2]:
								actionList.insert(sNum,2)
							if t > ultDic2[hero][2]:
								ultDic2[hero][3]=0
								ultDic2[hero][1]=0
								ultDic2[hero][2]=0

		if action =='Ult used':
			playernum=int(subject[6:len(subject)])
			st=BATTLE.t2s(time)

			if playernum <=6:
				ultDic1[hero][3]=playernum
				ultDic1[hero][1]=st
				ultDic1[hero][2]=st+ultDic1[hero][0]
			else:			
				ultDic2[hero][3]=playernum
				ultDic2[hero][1]=st
				ultDic2[hero][2]=st+ultDic2[hero][0]

			actionList.insert(sNum,actionNum[action])

		if action !='Hero switch' and action !='Eliminate':
			actionList.insert(sNum,actionNum[action])


		#队员
		if suicide ==1:
			subjectList.insert(sNum,int(obj[6:len(obj)]))
			suicide=0
		else:
			if subject != None:		
				subjectList.insert(sNum,int(subject[6:len(subject)]))
			else:
				subjectList.insert(sNum,subject)

		if obj !=None:
			objectList.insert(sNum,int(obj[6:len(obj)]))
		else :
			objectList.insert(sNum,obj)

		#时间
		timeList.insert(sNum,time)

		sNum+=1
		r+=1

	for j in range(sNum,sNum+5):
		timeList.insert(j,'59:59:59')
		actionList.insert(j,0)
		subjectList.insert(j,0)
		objectList.insert(j,0)

	#输入比赛信息
	if MATCH.attack == 'A':
		attack=MATCH.teama
	if MATCH.attack == 'B':
		attack=MATCH.teamb
	sheet_calcu.cell(row=2,column=2).value=attack
	sheet_calcu.cell(row=1,column=4).value=MATCH.teama
	sheet_calcu.cell(row=2,column=4).value=MATCH.teamb
	sheet_calcu.cell(row=2,column=8).value=mapdic[MATCH.mapn]
	gametime=0
	for i in range(0,len(ROUND.startlist)):
		gametime+=ROUND.t2s(ROUND.endlist[i])-ROUND.t2s(ROUND.startlist[i])
	m,s=divmod(gametime,60)
	h,m=divmod(m,60)
	print ("%02d:%02d:%02d" % (h, m, s))
	sheet_calcu.cell(row=2,column=10).value=("%02d:%02d:%02d" % (h, m, s))

	if MATCH.finalwinner == 'A':
		finalwinner=MATCH.teama
	if MATCH.finalwinner =='B':
		finalwinner=MATCH.teamb
	sheet_calcu.cell(row=2,column=11).value=finalwinner

	TEAM=Teamdic[MATCH.teama]
	r=23
	for i in range(0,6):
		sheet_calcu.cell(row=r+i,column=2).value=TEAM[MATCH.Aplayerlist[i]]
		sheet_calcu.cell(row=r+i,column=3).value=herodic[MATCH.Aherolist[i]]

	TEAM=Teamdic[MATCH.teamb]
	r=29
	for i in range(0,6):
		sheet_calcu.cell(row=r+i,column=2).value=TEAM[MATCH.Bplayerlist[i]]
		sheet_calcu.cell(row=r+i,column=3).value=herodic[MATCH.Bherolist[i]]

	c=12
	n=len(ROUND.startlist)
	for i in range(0,n):
		sheet_calcu.cell(row=75,column=c+i).value=ROUND.startlist[i]

	sn=0
	interval=0
	battle_s=0
	battle_e=0
	scene_s=0
	scene_e=0
	next_battle=1
	err=0
	
	xjbchange=[999,999,999,999,999,999,999,999,999,999,999,999]

	while scene_s < scount:
		if BATTLE.t2s(timeList[sn])<=BATTLE.t2s(sstartlist[scene_s]):
			if int(actionList[sn]) <300 and int(actionList[sn]) >200:
				xjbchange[subjectList[sn]-1]=int(actionList[sn])				

			timeList.pop(sn)
			actionList.pop(sn)
			subjectList.pop(sn)
			objectList.pop(sn)

		else:
			if interval !=0:
				timeList.insert(sn,sstartlist[scene_s])
				actionList.insert(sn,16)
				subjectList.insert(sn,None)
				objectList.insert(sn,None)

			sn+=1
			for player in range(0,12): 
				if xjbchange[player] !=999:
					timeList.insert(sn,sstartlist[scene_s])
					actionList.insert(sn,xjbchange[player])
					subjectList.insert(sn,player+1)
					objectList.insert(sn,None)
					xjbchange[player]=999
					sn+=1

			while BATTLE.t2s(timeList[sn]) <= BATTLE.t2s(sendlist[scene_s]):
				
				if battle_s < count and next_battle==1:
					if BATTLE.t2s(timeList[sn])<=BATTLE.t2s(startlist[battle_s]):
						sn+=1
						continue
					else:
						timeList.insert(sn,startlist[battle_s])
						actionList.insert(sn,8) #插入开团事件
						subjectList.insert(sn,None)
						objectList.insert(sn,None)
						battle_s+=1
						next_battle=0

				if battle_e < count:
					if BATTLE.t2s(endlist[battle_e]) >= BATTLE.t2s(sendlist[scene_s]):
						err=1
						err_battle=battle_e
					if BATTLE.t2s(timeList[sn]) <= BATTLE.t2s(endlist[battle_e]):
						sn+=1
						continue
					else:
						if err!=1:	
							timeList.insert(sn,endlist[battle_e])
							actionList.insert(sn,9) #插入关团事件
							subjectList.insert(sn,None)
							objectList.insert(sn,None)
							sn+=1
							timeList.insert(sn,endlist[battle_e])
							actionList.insert(sn,int(winlist[battle_e])) #插入团战胜利方
							subjectList.insert(sn,None)
							objectList.insert(sn,None)
							battle_e+=1
							next_battle=1
				sn+=1

			else:
				if err ==1:
					timeList.insert(sn,endlist[err_battle])
					actionList.insert(sn,9) #插入关团事件
					subjectList.insert(sn,None)
					objectList.insert(sn,None)
					sn+=1
					timeList.insert(sn,endlist[err_battle])
					actionList.insert(sn,int(winlist[err_battle])) #插入团战胜利方
					subjectList.insert(sn,None)
					objectList.insert(sn,None)
					battle_e+=1
					next_battle=1
					sn+=1
					err=0

				timeList.insert(sn,sendlist[scene_s])
				actionList.insert(sn,15)
				subjectList.insert(sn,None)
				objectList.insert(sn,None)
				sn+=1
				interval=sn
				scene_s+=1
	else:
		timeList.insert(sn,'0:00:00')
		actionList.insert(sn,16)
		subjectList.insert(sn,None)
		objectList.insert(sn,None)
		n=len(timeList)
		for i in range(sn+1,n):
			timeList.pop(sn+1)
			actionList.pop(sn+1)
			subjectList.pop(sn+1)
			objectList.pop(sn+1)
	
	r=79
	for i in range(0,len(timeList)):
		sheet_calcu.cell(row=r,column=3).value=subjectList[i]
		sheet_calcu.cell(row=r,column=7).value=actionList[i]
		sheet_calcu.cell(row=r,column=9).value=objectList[i]
		sheet_calcu.cell(row=r,column=14).value=timeList[i]
		r+=1	
	
	wb_calcu.save(filename=calcu_path)