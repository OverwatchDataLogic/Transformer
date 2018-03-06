#coding=utf-8

from Tkinter import (Tk,Message,Button,Entry,StringVar,Label,LabelFrame,
					Frame,Radiobutton,IntVar,Scrollbar,Listbox)
import tkFileDialog as tkf 
import tkMessageBox as tkb 
import ttk
from data import (RoundData,BattleData,MatchData,Path)
from openpyxl import Workbook
from openpyxl import load_workbook
import t2x


mapList=('Blizzard World', 'Dorado', 'Eichenwalde', 'Hanamura', 'Hollywood', 
		'Horizon Lunar Colony', 'Ilios', 'Junkertown', "King's Row",
	 	'Lijiang Tower', 'Nepal', 'Numbani', 'Oasis', 
	 	'Route 66', 'Temple of Anubis', 'Volskaya Industries', 'Watchpoint: Gibraltar')

heroList=('Ana', 'Bastion', 'D.Va', 'Doomfist', 'Genji', 
	'Hanzo', 'Junkrat', 'Lucio', 'Mccree', 'Mei', 'Mercy', 
	'Moira', 'Orisa', 'Pharah', 'Reaper', 'Reinhardt', 
	'Roadhog', 'Soldier: 76', 'Sombra', 'Symmetra', 'Torbjorn', 'Tracer',
 	'Widowmaker', 'Winston', 'Zarya', 'Zenyatta')
teamList=('BOS', 'DAL', 'FLA', 'GLA', 'HOU', 'LDN', 
		'NYE', 'PHI', 'SEO', 'SFS', 'SHD', 'VAL')

DALplayers=('AKM','CHIPSHAJEN', 'COCCO', 'CUSTA', 'EFFECT', 
			'HARRYHOOK', 'MICKIE', 'RASCAL','SEAGULL', 'TAIMOU', 'XQC')
PHIplayers=('BOOMBOX', 'CARPE', 'DAYFLY', 'EQO', 'FRAGI', 'HOTBA', 
			'JOWMEISTER', 'NEPTUNO', 'POKO', 'SADO', 'SHADOWBURN', 'SNILLO')
HOUplayers=('BANI', 'BOINK', 'CLOCKWORK', 'COOLMATT', 'JAKE',
			'LINKZR', 'MENDOKUSAII', 'MUMA', 'RAWKUS', 'SPREE')
SHDplayers=('ADO','ALTERING', 'DIYA', 'FEARLESS','FIVEKING', 'FREEFEEL', 'GEGURI'
			'MG', 'ROSHAN', 'SKY','UNDEAD', 'XUSHU')
BOSplayers=('AVAST', 'DREAMKAZPE', 'GAMSU', 'KALIOS',
	 		'KELLEX', 'MISTAKES', 'NEKO', 'NOTE', 'SNOW', 'STRIKER')
FLAplayers=('BISCHU', 'CWOOSH', 'LOGIX', 'MANNETEN', 'TVIQ', 'ZAPPIS','ZEBBOSAI', 'ZUPPEH')
NYEplayers=('ARK', 'JANUS', 'JJONAK', 'LIBERO', 'MANO', 'MEKO', 'PINE', 'SAEBYEOLBE')
SFSplayers=('BABYBAY', 'DANTEH', 'DHAK', 'IDDQD', 'NEVIX', 
			'NOMY', 'SINATRAA', 'SLEEPY', 'SUPER')
VALplayers=('AGILITIES', 'ENVY', 'FATE', 'GRIMREALITY', 'KARIV', 'NUMLOCKED',
			 'SILKTHREAD', 'SOON', 'SPACE', 'UNKOE', 'VERBO')
GLAplayers=('ASHER', 'BIGGOOSE', 'BISCHU', 'FISSURE','HYDRATION',
			 'IREMIIX', 'SHAZ', 'SUREFOUR')
SEOplayers=('BUNNY', 'FLETA', 'GIDO', 'KUKI', 'MIRO', 'MUNCHKIN', 
			'RYUJEHONG', 'TOBI', 'WEKEED', 'XEPHER', 'ZUNBA')
LDNplayers=('BDOSIN', 'BIRDRING', 'CLOSER', 'FURY', 'GESTURE', 
			'HAGOPEUN', 'HOOREG', 'NUS', 'PROFIT', 'WOOHYAL')

allPlayers={'DAL':DALplayers,'PHI':PHIplayers,'HOU':HOUplayers,'BOS':BOSplayers,
			'SHD':SHDplayers,'FLA':FLAplayers,'NYE':NYEplayers,'SFS':SFSplayers,
			'VAL':VALplayers,'GLA':GLAplayers,'SEO':SEOplayers,'LDN':LDNplayers}

class Display(object):
	def __init__(self):
		self.wb=load_workbook('save/save.xlsx',read_only=False)

		self.window=Tk()
		self.window.title("Transformer")
		self.window.geometry('960x600')

		self.ap=[]
		self.ah=[]
		self.bp=[]
		self.bh=[]

		self.MATCH=MatchData()
		self.BATTLE=BattleData()
		self.ROUND=RoundData()
		self.PATH=Path()

		self.ora_path=StringVar()
		self.calcu_path=StringVar()

		self.BATTLE.caption=StringVar()
		self.BATTLE.caption.set('Battle 1')
		self.BATTLE.end_time=StringVar()
		self.BATTLE.end_time.set('0:00:00')
		self.BATTLE.start_time=StringVar()
		self.BATTLE.start_time.set('0:00:00')

		self.ROUND.caption=StringVar()
		self.ROUND.caption.set('ROUND 1')
		self.ROUND.end_time=StringVar()
		self.ROUND.end_time.set('0:00:00')
		self.ROUND.start_time=StringVar()
		self.ROUND.start_time.set('0:00:00')

		self.MATCH.gameNum=StringVar()
		self.MATCH.mapNum=StringVar()
		self.MATCH.teamA=StringVar()
		self.MATCH.teamB=StringVar()
		self.MATCH.gameTime=StringVar()

		self.mode()
		self.path()
		self.inputinfo()
		self.showinfo()
		self.run()

	#Logic function
	def selectPath1(self):
		path_=tkf.askopenfilename()
		self.ora_path.set(path_)
	def selectPath2(self):
		path_=tkf.askopenfilename()
		self.calcu_path.set(path_)

	def loadwork(self):
		sheet=self.wb.worksheets[0]

		self.MATCH.game=sheet.cell(row=1,column=2).value
		self.MATCH.mapn=sheet.cell(row=1,column=4).value
		self.MATCH.teama=sheet.cell(row=3,column=2).value
		self.MATCH.teamb=sheet.cell(row=3,column=4).value
		self.MATCH.time=sheet.cell(row=14,column=2).value

		self.MATCH.gameNum.set(str(self.MATCH.game)) #场次
		self.MATCH.mapNum.set(self.MATCH.mapn)
		self.MATCH.teamA.set(self.MATCH.teama) #a队名
		self.MATCH.teamB.set(self.MATCH.teamb) #b队名

		self.MATCH.attack=sheet.cell(row=12,column=2).value

		self.MATCH.gameTime.set(str(self.MATCH.time))
		self.MATCH.finalwinner=sheet.cell(row=14,column=4).value

		self.lb.insert('end','Game'+str(self.num_entry.get()))
		self.lb.insert('end','Map: '+ str(self.mapName_cb.get()))		
		self.lb.insert('end',' ')

		for i in range(0,6):
			aplayer=sheet.cell(row=5+i,column=1).value #a队首发
			ahero=sheet.cell(row=5+i,column=2).value #a队首发英雄
			bplayer=sheet.cell(row=5+i,column=3).value #b队首发
			bhero=sheet.cell(row=5+i,column=4).value #b队首发英雄

			self.MATCH.Aplayerlist.insert(i,aplayer)
			self.MATCH.Aherolist.insert(i,ahero)
			self.MATCH.Bplayerlist.insert(i,bplayer)
			self.MATCH.Bherolist.insert(i,bhero)
			i+=1

		self.lb.insert('end','TeamA: '+ self.A.get())
		for i in range(0,6):
			self.lb.insert('end',self.MATCH.Aplayerlist[i]+':'+self.MATCH.Aherolist[i])
		self.lb.insert('end',' ')

		self.lb.insert('end','TeamB: '+ self.B.get())
		for i in range(0,6):
			self.lb.insert('end',self.MATCH.Bplayerlist[i]+':'+self.MATCH.Bherolist[i])
		self.lb.insert('end',' ')

		self.lb.insert('end','First Attack Team: '+ self.MATCH.attack)
		self.lb.insert('end','Game Duration: '+ self.time_entry.get())
		self.lb.insert('end','Game Winner: '+ self.MATCH.finalwinner)

		self.lb.insert('end','================#FIGHT DATA#================')
		r=3
		i=0
		while 1: #载入团战数据
			st=str(sheet.cell(row=r,column=6).value)
			et=str(sheet.cell(row=r,column=7).value)
			w=sheet.cell(row=r,column=8).value

			if st == 'None':
				break

			self.BATTLE.startlist.insert(i,st)
			self.BATTLE.endlist.insert(i,et)
			self.BATTLE.winlist.insert(i,w)

			if w==401:
				winteam=str(self.MATCH.teama)
			if w==402:
				winteam=str(self.MATCH.teamb)
			i+=1
			r+=1
			self.lb.insert('end','Fight'+str(i)+' : '+st+' - '+et+'   Winner:'+winteam)

		if len(self.BATTLE.startlist) > i:
			del self.BATTLE.startlist[i:len(self.BATTLE.startlist)]
			del self.BATTLE.endlist[i:len(self.BATTLE.endlist)]
			del self.BATTLE.winlist[i:len(self.BATTLE.winlist)]

		if i!=0:
			self.BATTLE.count=i
			self.BATTLE.count_=0
			self.BATTLE.temp1='0:00:00' #暂存团战的开始结束时间
			self.BATTLE.temp2='0:00:00'
			self.BATTLE.temp_old=str(self.BATTLE.endlist[self.BATTLE.count-2] )#上一波团战结束时间
			self.BATTLE.temp_next='59:59:59' #下一波团战开始时间
			self.BATTLE.caption.set('Fight %d'%(self.BATTLE.count))

		self.lb.insert('end','================#ROUND DATA#================')


		c=12
		i=0
		while 1: #载入场景时间
			st=str(sheet.cell(row=1,column=c).value)
			et=str(sheet.cell(row=2,column=c).value)

			if st=='None':
				break

			self.ROUND.startlist.insert(i,st)
			self.ROUND.endlist.insert(i,et)
			i+=1
			c+=1
			self.lb.insert('end','ROUND'+str(i)+' : '+st+' - '+et)	

		if len(self.ROUND.startlist) > i:
			del self.ROUND.startlist[i:len(self.ROUND.startlist)]
			del self.ROUND.endlist[i:len(self.ROUND.endlist)]

		if i!=0:
			self.ROUND.count=i
			self.ROUND.count_=0
			self.ROUND.temp1='0:00:00' #暂存场景的开始结束时间
			self.ROUND.temp2='0:00:00'
			self.ROUND.temp_old=str(self.ROUND.endlist[self.ROUND.count-2])#上一波场景结束时间
			self.ROUND.temp_next='59:59:59' #下一波场景开始时间
			self.ROUND.caption.set('Round %d'%(self.ROUND.count))

		tkb.showinfo(message='Load Success!')

	def newwork(self):		
		sheet=self.wb.worksheets[0]

		sheet.cell(row=1,column=2).value=None #场次
		sheet.cell(row=1,column=4).value=None #地图

		sheet.cell(row=3,column=2).value=None #a队名
		sheet.cell(row=3,column=4).value=None #b队名

		sheet.cell(row=12,column=2).value=None 

		sheet.cell(row=14,column=2).value=None
		sheet.cell(row=14,column=4).value=None

		for i in range(0,6):
			sheet.cell(row=5+i,column=1).value=None #a队首发
			sheet.cell(row=5+i,column=2).value=None #a队首发英雄
			sheet.cell(row=5+i,column=3).value=None #b队首发
			sheet.cell(row=5+i,column=4).value=None #b队首发英雄
			i+=1

		r=3
		while 1: #清空团战数据
			if sheet.cell(row=r,column=6).value==None:
				break
			sheet.cell(row=r,column=6).value=None 
			sheet.cell(row=r,column=7).value=None 
			sheet.cell(row=r,column=8).value=None
			r+=1
		
		c=12
		while 1:
			if sheet.cell(row=1,column=c).value==None:
				break
			sheet.cell(row=1,column=c).value=None 
			sheet.cell(row=2,column=c).value=None
			c+=1	

		#团战信息变量=====================
		self.BATTLE.init()
		self.BATTLE.caption.set('Fight 1')
		self.BATTLE.start_time.set('0:00:00')	
		self.BATTLE.end_time.set('0:00:00')

		#场景信息变量=======================
		self.ROUND.init()
		
		self.ROUND.start_time.set('0:00:00')	
		self.ROUND.end_time.set('0:00:00')
		
		self.ROUND.caption.set('Round 1')

		self.MATCH.init()
		self.MATCH.gameNum.set(' ')
		self.MATCH.mapNum.set(' ')
		self.MATCH.teamA.set(' ')
		self.MATCH.teamB.set(' ')
		self.wb.save(filename='save/save.xlsx')

		tkb.showinfo(message='New Work Created！')		

	def savematch(self):
		sheet=self.wb.worksheets[0]

		game=self.num_entry.get()
		mapn=self.mapName_cb.get()
		time=self.time_entry.get()

		sheet.cell(row=1,column=2).value=game #场次
		self.MATCH.game=game
		sheet.cell(row=1,column=4).value=mapn #地图
		self.MATCH.mapn=mapn
		sheet.cell(row=14,column=2).value=time
		self.MATCH.time=time

		teama=self.A.get()
		teamb=self.B.get()

		sheet.cell(row=3,column=2).value=teama #a队名
		self.MATCH.teama=teama
		sheet.cell(row=3,column=4).value=teamb #b队名
		self.MATCH.teamb=teamb

		for i in range(0,6):
			aplayer=self.ap[i].get()
			ahero=self.ah[i].get()
			bplayer=self.bp[i].get()
			bhero=self.bh[i].get()

			sheet.cell(row=5+i,column=1).value=aplayer #a队首发
			sheet.cell(row=5+i,column=2).value=ahero #a队首发英雄
			sheet.cell(row=5+i,column=3).value=bplayer #b队首发
			sheet.cell(row=5+i,column=4).value=bhero #b队首发英雄

			self.MATCH.Aplayerlist.insert(i,aplayer)
			self.MATCH.Aherolist.insert(i,ahero)
			self.MATCH.Bplayerlist.insert(i,bplayer)
			self.MATCH.Bherolist.insert(i,bhero)

			i+=1

		win=self.winteam.get()
		if win ==1:
			self.MATCH.finalwinner='A'
			sheet.cell(row=14,column=4).value='A'
		else :
			self.MATCH.finalwinner='B'
			sheet.cell(row=14,column=4).value='B'

		attack=self.attackteam.get()
		if attack == 1:
			self.MATCH.attack='A'
			sheet.cell(row=12,column=2).value='A'
		else :
			self.MATCH.attack='B'
			sheet.cell(row=12,column=2).value='B'

		self.wb.save(filename='save/save.xlsx')

		tkb.showinfo(message='Match infomation saved')		

	def battlenext(self):
		count=self.BATTLE.count
		count_=self.BATTLE.count_
		temp_old=self.BATTLE.temp_old
		temp_next=self.BATTLE.temp_next
		startlist=self.BATTLE.startlist
		endlist=self.BATTLE.endlist
		winlist=self.BATTLE.winlist

		temp1=self.start_entry.get()
		temp2=self.end_entry.get()
		if self.BATTLE.checkformat(temp1) ==False:		#检验输入格式
			tkb.showwarning(message='Format of start-time is wrong, please re-input(e.g.0:01:00)')
			return 
		if self.BATTLE.checkformat(temp2) ==False:		#检验输入格式
			tkb.showwarning(message='Format of end-time is wrong, please re-input(e.g.0:01:00)')
			return 

		if count_ !=0 and count!=count_-1:
			temp_next=startlist[count]
			if self.BATTLE.t2s(temp2) >=self.BATTLE.t2s(temp_next):
				tkb.showwarning(message='The end time is later than the start time of the next fight，please re-input')
				return

		if count_ !=0 and count == count_-1:
			temp_old=endlist[count_-2]
			if self.BATTLE.t2s(temp1) <= self.BATTLE.t2s(temp_old):		#逻辑检验
				tkb.showwarning(message='The start time is earlier than the end time of the last fight，please re-input')
				return 

		if count_ ==0 and count > 1:
			temp_old=endlist[count-2]
			if self.BATTLE.t2s(temp1) <= self.BATTLE.t2s(temp_old):		#逻辑检验
				tkb.showwarning(message='The start time is earlier than the end time of the last fight，please re-input')
				return 	

		if self.BATTLE.t2s(temp1) >= self.BATTLE.t2s(temp2):		#逻辑检验
			tkb.showwarning(message='The end time is earlier than the start time in a fight，please re-input') 
			return 	

		self.BATTLE.winteam=self.team.get()

		startlist.insert(count-1,temp1)
		endlist.insert(count-1,temp2)
		winlist.insert(count-1,self.BATTLE.winteam+400)
		if self.BATTLE.winteam==1:
			winteam=self.A.get()
		if self.BATTLE.winteam==2:
			winteam=self.B.get()

		self.lb.insert('end','Fight'+str(count)+' : '+startlist[count-1]+' - '+endlist[count-1]+'   Winner:'+ winteam)

		if count_ !=0:
			count=count_
			count_=0
		else:
			count+=1

		self.BATTLE.count=count
		self.BATTLE.count_=count_
		self.BATTLE.temp_old=temp_old
		self.BATTLE.temp_next=temp_next
		self.BATTLE.startlist=startlist
		self.BATTLE.endlist=endlist
		self.BATTLE.winlist=winlist

		self.BATTLE.caption.set('Fight %d'%(count))
		#start_time.set("")
		#end_time.set("")

	def fightalter(self):
		count=self.BATTLE.count
		count_=self.BATTLE.count_

		if count <=1:
			return 
		count_=count
		count=int(self.alter_entry.get())
		if count>=count_:
			tkb.showwarning(message='Data of this fight has not been inputted, please re-input the number')
			count=count_
			count_=0
			return

		self.lb.insert('end','Ready to alter fight '+str(count))
		self.BATTLE.caption.set('Fight %d'%(count))	

		self.BATTLE.count=count
		self.BATTLE.count_=count_

	def savefight(self):
		sheet=self.wb.worksheets[0]

		r=3
		while 1: #清空团战数据
			if sheet.cell(row=r,column=6).value==None:
				break
			sheet.cell(row=r,column=6).value=None 
			sheet.cell(row=r,column=7).value=None 
			sheet.cell(row=r,column=8).value=None
			r+=1
		
		n=len(self.BATTLE.startlist)

		for i in range(0,n):
			sheet.cell(row=r,column=6).value=self.BATTLE.startlist[i]
			sheet.cell(row=r,column=7).value=self.BATTLE.endlist[i]
			sheet.cell(row=r,column=8).value=self.BATTLE.winlist[i]

		self.wb.save(filename='save/save.xlsx')	

	def roundnext(self):
		scount=self.ROUND.count
		scount_=self.ROUND.count_
		stemp_old=self.ROUND.temp_old
		stemp_next=self.ROUND.temp_next
		sstartlist=self.ROUND.startlist
		sendlist=self.ROUND.endlist

		stemp1=self.sstart_entry.get()
		stemp2=self.send_entry.get()
		if self.ROUND.checkformat(stemp1) ==False:		#检验输入格式
			tkb.showwarning(message='Format of start-time is wrong, please re-input(e.g.0:01:00)')
			return 
		if self.ROUND.checkformat(stemp2) ==False:		#检验输入格式
			tkb.showwarning(message='Format of end-time is wrong, please re-input(e.g.0:01:00)')
			return 

		if scount_ !=0 and scount!=scount_-1:
			stemp_next=sstartlist[scount]
			if self.ROUND.t2s(stemp2) >=self.ROUND.t2s(stemp_next):
				tkb.showwarning(message='The end time is later than the start time of the next round，please re-input')
				return

		if scount_ !=0 and scount == scount_-1:
			stemp_old=sendlist[scount_-2]
			if self.ROUND.t2s(stemp1) <= self.ROUND.t2s(stemp_old):		#逻辑检验
				tkb.showwarning(message='The start time is earlier than the end time of the last round，please re-input')
				return 

		if scount_ ==0 and scount > 1:
			stemp_old=sendlist[scount-2]
			if self.ROUND.t2s(stemp1) <= self.ROUND.t2s(stemp_old):		#逻辑检验
				tkb.showwarning(message='The start time is earlier than the end time of the last round，please re-input')
				return 	

		if self.ROUND.t2s(stemp1) >= self.ROUND.t2s(stemp2):		#逻辑检验
			tkb.showwarning(message='The end time is earlier than the start time in a round，please re-input') 
			return 	

		sstartlist.insert(scount-1,stemp1)
		sendlist.insert(scount-1,stemp2)

		self.lb.insert('end','Round'+str(scount)+' : '+sstartlist[scount-1]+' - '+sendlist[scount-1])

		if scount_ !=0:
			scount=scount_
			scount_=0
		else:
			scount+=1

		self.ROUND.caption.set('Round %d'%(scount))

		self.ROUND.count=scount
		self.ROUND.count_=scount_
		self.ROUND.temp_old=stemp_old
		self.ROUND.temp_next=stemp_next
		self.ROUND.startlist=sstartlist
		self.ROUND.endlist=sendlist

	def roundalter(self):
		scount=self.ROUND.count
		scount=self.ROUND.count_

		if scount <=1:
			return 
		scount_=scount
		scount=int(entry_smodi.get())
		if scount>=scount_:
			tkb.showwarning(message='Data of this round has not been inputted, please re-input the number')
			scount=scount_
			scount_=0
			return

		self.lb.insert('end','Ready to alter Round '+str(scount))
		self.ROUND.caption.set('Round %d'%(scount))

		self.ROUND.count=scount
		self.ROUND.count_=scount

	def saveround(self):
		sheet=self.wb.worksheets[0]

		c=12
		while 1:
			if sheet.cell(row=1,column=c).value==None:
				break
			sheet.cell(row=1,column=c).value=None 
			sheet.cell(row=2,column=c).value=None
			c+=1	

		n=len(self.ROUND.startlist)

		for i in range(0,n):
			sheet.cell(row=1,column=c).value=self.ROUND.startlist[i] 
			sheet.cell(row=2,column=c).value=self.ROUND.endlist[i]

		self.wb.save(filename='save/save.xlsx')

	def trans(self):
		orapath=self.ora_entry.get()
		calcupath=self.calcu_entry.get()
		t2x.trans(self.MATCH,self.BATTLE,self.ROUND,orapath,calcupath)
		tkb.showinfo(message='Completed!')

	#UI
	def mode(self):
		load=Button(self.window,text='Load',width=10,command=self.loadwork)
		load.place(x=20,y=20)

		new=Button(self.window,text='New Work',width=10,command=self.newwork)
		new.place(x=120,y=20)

	def path(self):
		path_frame=LabelFrame(self.window,text='Input paths',width=450,height=100)
		path_frame.place(x=20,y=60)

		ora_mes=Message(path_frame,text='Excel from ORA:',width=120)
		ora_mes.place(x=10,y=0)
		self.ora_entry=Entry(path_frame,textvariable=self.ora_path,width=35)
		self.ora_entry.place(x=130,y=0)
		ora_button=Button(path_frame,text='Select…',command=self.selectPath1)
		ora_button.place(x=365,y=0)

		calcu_mes=Message(path_frame,text='Calculator:',width=120)
		calcu_mes.place(x=10,y=40)
		self.calcu_entry=Entry(path_frame,textvariable=self.calcu_path,width=35)
		self.calcu_entry.place(x=130,y=40)
		calcu_button=Button(path_frame,text='Select…',command=self.selectPath2)
		calcu_button.place(x=365,y=40)

	def inputinfo(self):
		tabControl=ttk.Notebook(self.window)

		match=ttk.Frame(tabControl,width=450,height=328)
		tabControl.add(match,text=' -Match- ')

		battle=ttk.Frame(tabControl,width=450,height=328)
		tabControl.add(battle,text=' -Fight- ')

		scene=ttk.Frame(tabControl,width=450,height=328)
		tabControl.add(scene,text=' -Round- ')

		tabControl.place(x=20,y=180)

		#Match Info
		num=Message(match,text='Game No.',width=70)
		num.place(x=10,y=10)

		self.num_entry=Entry(match,width=6,textvariable=self.MATCH.gameNum)
		self.num_entry.place(x=80,y=10)


		mapName=Message(match,text='Map:',width=50)
		mapName.place(x=130,y=10)

		self.mapName_cb=ttk.Combobox(match,width=10,textvariable=self.MATCH.mapNum)
		self.mapName_cb['values']=mapList
		self.mapName_cb.place(x=180,y=10)

		time=Message(match,text='Duration:',width=100)
		time.place(x=280,y=10)

		self.time_entry=Entry(match,width=10,textvariable=self.MATCH.gameTime)
		self.time_entry.place(x=350,y=10)

		def selectTeamA(*args):			
			team_selected=teamA_cb.current()
			y=20

			for i in range(0,6):
				self.ap.insert(i,ttk.Combobox(teamA_frame,width=10))
				print self.ap[i]
				self.ap[i]['values']=allPlayers.get(teamList[team_selected],' ')
				name='Player'+str(i+1)
				self.ap[i].insert(0,name)
				self.ap[i].place(x=10,y=y+20*(i+1))
			for i in range(0,6):
				self.ah.insert(i,ttk.Combobox(teamA_frame,width=10))
				self.ah[i]['values']=heroList
				name='hero'
				self.ah[i].insert(0,name)
				self.ah[i].place(x=105,y=y+20*(i+1))
		def selectTeamB(*args):
			team_selected=teamB_cb.current()
			y=20
			for i in range(0,6):
				self.bp.insert(i,ttk.Combobox(teamB_frame,width=10))
				self.bp[i]['values']=allPlayers.get(teamList[team_selected],' ')
				name='Player'+str(i+7)
				self.bp[i].insert(0,name)
				self.bp[i].place(x=10,y=y+20*(i+1))
			for i in range(0,6):
				self.bh.insert(i,ttk.Combobox(teamB_frame,width=10))
				self.bh[i]['values']=heroList
				name='hero'
				self.bh[i].insert(0,name)
				self.bh[i].place(x=105,y=y+20*(i+1))

		#Team A information
		teamA_frame=LabelFrame(match,text='TeamA',width=210,height=200)
		teamA_frame.place(x=10,y=50)

		teamA_name=Message(teamA_frame,text='Name:',width=70)
		teamA_name.place(x=10,y=10)

		teamA_cb=ttk.Combobox(teamA_frame,width=12,textvariable=self.MATCH.teamA)
		teamA_cb['values']=teamList
		teamA_cb.bind('<<ComboboxSelected>>',selectTeamA)
		teamA_cb.place(x=60,y=10)

		#Team B information
		teamB_frame=LabelFrame(match,text='TeamB',width=210,height=200)
		teamB_frame.place(x=230,y=50)

		teamB_name=Message(teamB_frame,text='Name:',width=70)
		teamB_name.place(x=10,y=10)

		teamB_cb=ttk.Combobox(teamB_frame,width=12,textvariable=self.MATCH.teamB)
		teamB_cb['values']=teamList
		teamB_cb.bind('<<ComboboxSelected>>',selectTeamB)
		teamB_cb.place(x=60,y=10)

		self.A=teamA_cb
		self.B=teamB_cb

		winner=Message(match,text='Match Winner:',width=150)
		winner.place(x=10,y=260)
		self.winteam=IntVar()
		if self.MATCH.finalwinner == 'A':
			self.winteam.set(1)
		if self.MATCH.finalwinner == 'B':
			self.winteam.set(2)
		winA=Radiobutton(match,text='TeamA',variable=self.winteam,value=1)
		winB=Radiobutton(match,text='TeamB',variable=self.winteam,value=2)
		winA.place(x=110,y=260)
		winB.place(x=180,y=260)

		attacker=Message(match,text='First Attack:',width=150)
		attacker.place(x=10,y=280)
		
		self.attackteam=IntVar()
		if self.MATCH.attack == 'A':
			self.attackteam.set(1)
		if self.MATCH.attack == 'B':
			self.attackteam.set(2)
		attackA=Radiobutton(match,text='TeamA',variable=self.attackteam,value=1)
		attackB=Radiobutton(match,text='TeamB',variable=self.attackteam,value=2)
		attackA.place(x=110,y=280)
		attackB.place(x=180,y=280)

		matchSave=Button(match,text='Save',width=10,command=self.savematch)
		matchSave.place(x=330,y=290)

		#Battle Info
		battle_label=Label(battle,textvariable=self.BATTLE.caption,width=15,height=1)
		battle_label.place(x=0,y=20)

		start_tag=Label(battle,text='Start time:',foreground='black',width=8,height=1)
		start_tag.place(x=100,y=10)

		self.start_entry=Entry(battle,width=20,textvariable=self.BATTLE.start_time,foreground='black')
		self.start_entry.place(x=200,y=10)

		end_tag=Label(battle,text='End time:',foreground='black',width=8,height=1)
		end_tag.place(x=100,y=40)

		self.end_entry=Entry(battle,width=20,textvariable=self.BATTLE.end_time,foreground='black')
		self.end_entry.place(x=200,y=40)

		#winner
		win_tag=Label(battle,text='Fight Winner:',foreground='black',width=10,height=1) #标签
		win_tag.place(x=100,y=70)

		self.team=IntVar()
		self.team.set(1)
		team1=Radiobutton(battle,text='Team A',variable=self.team,value=1)
		team1.place(x=200,y=70)
		team2=Radiobutton(battle,text='Team B',variable=self.team,value=2)
		team2.place(x=270,y=70)

		next_button=Button(battle,text='Next',width=10,command=self.battlenext)
		next_button.place(x=100,y=100)

		alter_tag=Label(battle,text='Input fight number to alter information:') #标签
		alter_tag.place(x=30,y=150)

		self.alter_entry=Entry(battle,width=5)
		self.alter_entry.place(x=280,y=150)

		alter_button=Button(battle,text='Go',width=7,command=self.fightalter)
		alter_button.place(x=330,y=150)

		battleSave=Button(battle,text='Save',width=10,command=self.savefight)
		battleSave.place(x=330,y=290)

		#Round Info
		round_label=Label(scene,textvariable=self.ROUND.caption,width=15,height=1)
		round_label.place(x=0,y=20)

		sstart_tag=Label(scene,text='Start time:',foreground='black',width=8,height=1)
		sstart_tag.place(x=100,y=10)

		self.sstart_entry=Entry(scene,width=20,textvariable=self.ROUND.start_time,foreground='black')
		self.sstart_entry.place(x=200,y=10)

		send_tag=Label(scene,text='End time:',foreground='black',width=8,height=1)
		send_tag.place(x=100,y=40)

		self.send_entry=Entry(scene,width=20,textvariable=self.ROUND.end_time,foreground='black')
		self.send_entry.place(x=200,y=40)

		snext_button=Button(scene,text='Next',width=10,command=self.roundnext)
		snext_button.place(x=100,y=100)

		salter_tag=Label(scene,text='Input battle number to alter information:') #标签
		salter_tag.place(x=30,y=150)

		self.salter_entry=Entry(scene,width=5)
		self.salter_entry.place(x=280,y=150)

		salter_button=Button(scene,text='Go',width=7,command=self.roundalter)
		salter_button.place(x=330,y=150)

		sceneSave=Button(scene,text='Save',width=10,command=self.saveround)
		sceneSave.place(x=330,y=290)

	def showinfo(self):
		infoRecord=LabelFrame(self.window,text='Operation Note')
		infoRecord.place(x=485,y=60)

		sl=Scrollbar(infoRecord)
		sl.pack(side='right',fill='y')
		lb=Listbox(infoRecord,height=25,width=60)
		lb.pack()

		lb['yscrollcommand']=sl.set
		sl['command']=lb.yview

		self.lb=lb

	def run(self):
		run_button=Button(self.window,text='Run',width=15,command=self.trans)
		run_button.place(x=810,y=550)

	def show(self):
		self.window.mainloop()





